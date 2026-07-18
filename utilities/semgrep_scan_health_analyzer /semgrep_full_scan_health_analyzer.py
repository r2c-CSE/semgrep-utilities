import requests
import csv
from typing import List, Dict
import os
from datetime import datetime
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

class SemgrepAPI:
    BASE_URL = "https://semgrep.dev/api/v1"
    
    def __init__(self, api_token: str, org_id: str, deployment_slug: str):
        self.api_token = api_token
        self.org_id = org_id
        self.deployment_slug = deployment_slug
        self.headers = {
            "Authorization": f"Bearer {api_token}",
            "Content-Type": "application/json"
        }

    def get_projects(self) -> List[Dict]:
        """Get all projects in the organization using pagination"""
        all_projects = []
        page = 0
        page_size = 3000  # Maximum allowed by API

        while True:
            url = f"{self.BASE_URL}/deployments/{self.deployment_slug}/projects"
            params = {
                "page": page,
                "page_size": page_size
            }
            
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            
            data = response.json()
            projects = data.get("projects", [])
            
            if not projects:
                break
                
            all_projects.extend(projects)
            page += 1
            
        return all_projects

    def get_scans_for_project(self, repository_id: str, project_name: str) -> List[Dict]:
        """Get all scans for a specific project"""
        print(f"Getting scans for project {repository_id}")
        url = f"{self.BASE_URL}/deployments/{self.org_id}/scans/search"
        
        # Send parameters in request body
        data = {
            "repository_id": int(repository_id),
            "is_full_scan": True,
            "limit": 0
        }
        
        response = requests.post(url, headers=self.headers, json=data)
        response.raise_for_status()
        
        print(f"Response: {response.json()}")

        # Add project_name to each scan
        scans = response.json().get("scans", [])
        for scan in scans:
            scan["project_name"] = project_name
            
        return scans


def write_scans_to_csv(scans: List[Dict], filename: str):
    """Write scan data to a CSV file"""
    if not scans:
        print("No scans found to write to CSV")
        return

    fieldnames = [
        "scan_id",
        "repository_id",
        "project_name",
        "is_full_scan",
        "status",
        "branch",
        "completed_at"
    ]

    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for scan in scans:
            row = {
                "scan_id": scan.get("id", ""),
                "repository_id": scan.get("repository_id", ""),
                "project_name": scan.get("project_name", ""),
                "is_full_scan": scan.get("is_full_scan", ""),
                "status": scan.get("status", ""),
                "branch": scan.get("branch", ""),
                "completed_at": scan.get("completed_at", "")
            }
            writer.writerow(row)

    print(f"Wrote {len(scans)} scans to CSV file: {filename}")

def write_scans_to_excel(wb: Workbook, scans: List[Dict]):
    """Write scan data to an Excel file with formatting"""
    if not scans:
        print("No scans found to write to Excel")
        return

    ws = wb.active
    ws.title = "Semgrep Scans"

    # Define headers and their widths
    headers = [
        "Scan ID",
        "Repository ID",
        "Project Name",
        "Is Full Scan",
        "Status",
        "Branch",
        "Completed At"
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Scan ID
    ws.column_dimensions['B'].width = 15  # Repository ID
    ws.column_dimensions['C'].width = 40  # Project Name
    ws.column_dimensions['D'].width = 15  # Is Full Scan
    ws.column_dimensions['E'].width = 20  # Status
    ws.column_dimensions['F'].width = 20  # Branch
    ws.column_dimensions['G'].width = 25  # Completed At

    # Write headers with formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data with proper type conversion
    for row, scan in enumerate(scans, 2):
        # Convert IDs to integers
        ws.cell(row=row, column=1, value=int(scan.get("id", 0)))
        ws.cell(row=row, column=2, value=int(scan.get("repository_id", 0)))
        ws.cell(row=row, column=3, value=scan.get("project_name", ""))
        ws.cell(row=row, column=4, value=scan.get("is_full_scan", ""))
        ws.cell(row=row, column=5, value=scan.get("status", ""))
        ws.cell(row=row, column=6, value=scan.get("branch", ""))
        ws.cell(row=row, column=7, value=scan.get("completed_at", ""))

    print(f"Wrote {len(scans)} scans to Excel sheet: Semgrep Scans")

def calculate_repo_health_metrics(scans: List[Dict]) -> List[Dict]:
    """Calculate health metrics for each repository"""
    repo_metrics = defaultdict(lambda: {"total_scans": 0, "failed_scans": 0, "project_name": ""})
    
    for scan in scans:
        repo_id = scan.get("repository_id")
        repo_metrics[repo_id]["total_scans"] += 1
        repo_metrics[repo_id]["project_name"] = scan.get("project_name", "")
        if scan.get("status") == "SCAN_STATUS_NEVER_FINISHED":
            repo_metrics[repo_id]["failed_scans"] += 1

    # Calculate failure percentages and create final metrics
    health_metrics = []
    for repo_id, metrics in repo_metrics.items():
        total_scans = metrics["total_scans"]
        failed_scans = metrics["failed_scans"]
        failure_percentage = (failed_scans / total_scans * 100) if total_scans > 0 else 0
        
        health_metrics.append({
            "repository_id": repo_id,
            "project_name": metrics["project_name"],
            "total_scans": total_scans,
            "failed_scans": failed_scans,
            "failure_percentage": round(failure_percentage, 2),
            "success_rate": round(100 - failure_percentage, 2)
        })
    
    # Sort by failure percentage in descending order
    return sorted(health_metrics, key=lambda x: x["failure_percentage"], reverse=True)

def write_health_metrics_to_csv(metrics: List[Dict], filename: str):
    """Write repository health metrics to a CSV file"""
    if not metrics:
        print("No metrics found to write to CSV")
        return

    fieldnames = [
        "repository_id",
        "project_name",
        "total_scans",
        "failed_scans",
        "failure_percentage",
        "success_rate"
    ]

    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(metrics)

    print(f"Wrote health metrics for {len(metrics)} repositories to CSV file: {filename}")

def write_health_metrics_to_excel(wb: Workbook, metrics: List[Dict]):
    """Add repository health metrics as a new sheet in the Excel workbook"""
    if not metrics:
        print("No metrics found to write to Excel")
        return

    ws = wb.create_sheet(title="Repo Health Summary")

    # Define headers and their widths
    headers = [
        "Repository ID",
        "Project Name",
        "Total Scans",
        "Failed Scans",
        "Failure %",
        "Success Rate %"
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Repository ID
    ws.column_dimensions['B'].width = 40  # Project Name
    ws.column_dimensions['C'].width = 15  # Total Scans
    ws.column_dimensions['D'].width = 15  # Failed Scans
    ws.column_dimensions['E'].width = 15  # Failure %
    ws.column_dimensions['F'].width = 15  # Success Rate

    # Write headers with formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data with conditional formatting and proper type conversion
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green

    for row, metric in enumerate(metrics, 2):
        # Convert numeric values appropriately
        ws.cell(row=row, column=1, value=int(metric["repository_id"]))
        ws.cell(row=row, column=2, value=metric["project_name"])
        ws.cell(row=row, column=3, value=int(metric["total_scans"]))
        ws.cell(row=row, column=4, value=int(metric["failed_scans"]))
        
        # Store percentages as numbers
        failure_cell = ws.cell(row=row, column=5, value=float(metric["failure_percentage"]))
        success_cell = ws.cell(row=row, column=6, value=float(metric["success_rate"]))
        
        # Format cells to display as percentages
        failure_cell.number_format = '0.00"%"'
        success_cell.number_format = '0.00"%"'
        
        # Color code based on failure percentage
        if metric["failure_percentage"] >= 20:
            failure_cell.fill = red_fill
            success_cell.fill = red_fill
        elif metric["failure_percentage"] >= 10:
            failure_cell.fill = yellow_fill
            success_cell.fill = yellow_fill
        else:
            failure_cell.fill = green_fill
            success_cell.fill = green_fill

def main():
    # Get inputs from environment variables
    api_token = os.getenv("SEMGREP_API_TOKEN")
    org_id = os.getenv("SEMGREP_ORG_ID")
    deployment_slug = os.getenv("SEMGREP_DEPLOYMENT_SLUG")

    # Validate required environment variables
    if not all([api_token, org_id, deployment_slug]):
        print("Error: Missing required environment variables")
        print("Please set the following environment variables:")
        print("  - SEMGREP_API_TOKEN")
        print("  - SEMGREP_ORG_ID")
        print("  - SEMGREP_DEPLOYMENT_SLUG")
        sys.exit(1)

    # Create reports directory if it doesn't exist
    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
        print(f"Created reports directory: {reports_dir}")

    # Initialize API client
    semgrep = SemgrepAPI(api_token, org_id, deployment_slug)

    # Get all projects
    print("Fetching projects...")
    projects = semgrep.get_projects()
    print(f"Found {len(projects)} projects")

    # Get scans for each project
    all_scans = []
    for project in projects:
        repository_id = project.get("id")
        project_name = project.get("name")
        print(f"Fetching scans for project: {project_name}")
        
        scans = semgrep.get_scans_for_project(repository_id, project_name)
        all_scans.extend(scans)
        print(f"Found {len(scans)} scans for project {project_name}")

    # Calculate repository health metrics
    health_metrics = calculate_repo_health_metrics(all_scans)

    # Write results to files
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Write scan details to CSV
    csv_filename = os.path.join(reports_dir, f"semgrep_scans_{timestamp}.csv")
    write_scans_to_csv(all_scans, csv_filename)
    
    # Create Excel workbook and write both scan details and health metrics
    excel_filename = os.path.join(reports_dir, f"semgrep_scans_{timestamp}.xlsx")
    wb = Workbook()
    
    # Write both sheets to the Excel workbook
    write_scans_to_excel(wb, all_scans)
    write_health_metrics_to_excel(wb, health_metrics)
    wb.save(excel_filename)
    
    # Write health metrics to separate CSV
    health_csv_filename = os.path.join(reports_dir, f"repo_scan_health_summary_{timestamp}.csv")
    write_health_metrics_to_csv(health_metrics, health_csv_filename)
    
    print(f"\nAll reports have been saved to the '{reports_dir}' directory:")

if __name__ == "__main__":
    main() 
